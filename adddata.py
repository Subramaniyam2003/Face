from tkinter import *
from tkinter import font
import openpyxl
import xlrd
import pathlib
from openpyxl import Workbook

q=Tk()
q.configure(bg='black')
q.title('ADD DATA')
q.geometry(("1600x556"))
def clear():
    for widget in [il1,na1,ge1,dob1,pn1,ad1]:
        widget.delete(0,'end')
    
def submit():
    a=il1.get()
    b=na1.get()
    c=ge1.get()
    d=dob1.get()
    e=pn1.get()
    f=ad1.get()
    print(a)
    print(b)
    print(c)
    print(d)
    print(e)
    print(f)
    file=openpyxl.load_workbook(r"C:\Users\OM MURUGA\Desktop\python\records.csv")
    sheet=file.active
    sheet.cell(column=1,row=sheet.max_row+1,value=a)
    sheet.cell(column=2,row=sheet.max_row+0,value=b)
    sheet.cell(column=3,row=sheet.max_row+0,value=c)
    sheet.cell(column=4,row=sheet.max_row+0,value=d)
    sheet.cell(column=5,row=sheet.max_row+0,value=e)
    sheet.cell(column=6,row=sheet.max_row+0,value=f)
    file.save(r"C:\Users\OM MURUGA\Desktop\python\records.csv")
    


    
file=pathlib.Path(r"C:\Users\OM MURUGA\Desktop\python\records.csv")
if file.exists():
    pass
else:
    file=Workbook()
    sheet=file.active
    sheet["A1"]="Image_Location"
    sheet["B1"]="Name"
    sheet["C1"]="Gender"
    sheet["D1"]="DOB"
    sheet["E1"]="Phone_No"
    sheet["F1"]="Address"
    file.save(r"C:\Users\OM MURUGA\Desktop\python\records.csv")
il=Label(q,text='ENTER IMAGE LOCATION  :',bg='black',fg='red')
il.place(x=100,y=80)
il.configure(font=( "Helvetica",14, "bold"))
il1=Entry(q,width=100)
il1.place(x=400,y=80)
na=Label(q,text='ENTER THE NAME        :',bg='black',fg='red')
na.place(x=100,y=160)
na.configure(font=( "Helvetica",14, "bold"))
na1=Entry(q,width=100)
na1.place(x=400,y=160)
ge=Label(q,text='ENTER THE GENDER      :',bg='black',fg='red')
ge.place(x=100,y=240)
ge.configure(font=( "Helvetica",14, "bold"))
ge1=Entry(q,width=100)
ge1.place(x=400,y=240)
dob=Label(q,text='ENTER THE DOB          :',bg='black',fg='red')
dob.place(x=100,y=320)
dob.configure(font=( "Helvetica",14, "bold"))
dob1=Entry(q,width=100)
dob1.place(x=400,y=320)
pn=Label(q,text='ENTER PHONE NUMBER    :',bg='black',fg='red')
pn.place(x=100,y=400)
pn.configure(font=( "Helvetica",14, "bold"))
pn1=Entry(q,width=100)
pn1.place(x=400,y=400)
ad=Label(q,text='ENTER THE ADDRESS     :',bg='black',fg='red')
ad.place(x=100,y=480)
ad.configure(font=( "Helvetica",14, "bold"))
ad1=Entry(q,width=100)
ad1.place(x=400,y=480)
b=Button(q,text='SUBMIT',fg='red',bg='black',activebackground='red',activeforeground='black',pady=5,command=submit)
b.place(x=700,y=570)
b.configure(font=( "Helvetica",14, "bold"))
b1=Button(q,text='CLEAR',fg='red',bg='black',activebackground='red',activeforeground='black',pady=5,command=clear)
b1.place(x=900,y=570)
b1.configure(font=( "Helvetica",14, "bold"))
mainloop()
